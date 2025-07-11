#!/usr/bin/env python3
"""
Extract Images from Excel File for Odu Cards
Extracts embedded images from the 256 Odu Excel matrix file
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import json
from datetime import datetime
from PIL import Image
import io

class ExcelImageExtractor:
    def __init__(self):
        self.excel_path = "attached_assets/256_ODU_graph_1752246322368.xlsx"
        self.output_dir = "static/odu_cards"
        self.images_dir = "static/odu_cards/images"
        
        # Create directories
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)
        
    def extract_images_from_excel(self):
        """Extract all images from Excel file"""
        try:
            print("Loading Excel workbook...")
            workbook = load_workbook(self.excel_path)
            worksheet = workbook.active
            
            print(f"Workbook loaded. Active sheet: {worksheet.title}")
            print(f"Sheet dimensions: {worksheet.max_row} x {worksheet.max_column}")
            
            # Check for images in the worksheet
            images = []
            
            # Method 1: Check for drawing objects
            if hasattr(worksheet, '_images'):
                print(f"Found {len(worksheet._images)} images in worksheet._images")
                for idx, image in enumerate(worksheet._images):
                    images.append(self.save_image_from_drawing(image, idx))
            
            # Method 2: Check for embedded objects
            if hasattr(worksheet, '_charts'):
                print(f"Found {len(worksheet._charts)} chart objects")
            
            # Method 3: Manual search through worksheet relations
            if worksheet._rels:
                print(f"Found {len(worksheet._rels)} relationships")
                for rel_id, rel in worksheet._rels.items():
                    print(f"Relationship: {rel_id} -> {rel.target}")
                    if 'image' in rel.target.lower() or 'media' in rel.target.lower():
                        try:
                            # Try to extract image data
                            image_data = self.extract_image_from_relation(workbook, rel)
                            if image_data:
                                images.append(image_data)
                        except Exception as e:
                            print(f"Error extracting image from relation {rel_id}: {e}")
            
            # Method 4: Check workbook-level images
            try:
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
                
                # Look for drawing parts in the workbook
                for part_name in workbook._archive.namelist():
                    if 'drawing' in part_name.lower() or 'image' in part_name.lower() or 'media' in part_name.lower():
                        print(f"Found potential image/drawing file: {part_name}")
                        
                        if part_name.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                            # This is an image file
                            image_data = workbook._archive.read(part_name)
                            image_filename = f"excel_image_{len(images)}.{part_name.split('.')[-1]}"
                            image_path = os.path.join(self.images_dir, image_filename)
                            
                            with open(image_path, 'wb') as f:
                                f.write(image_data)
                            
                            images.append({
                                'filename': image_filename,
                                'path': image_path,
                                'source': part_name,
                                'size': len(image_data)
                            })
                            
                            print(f"Extracted image: {image_filename} ({len(image_data)} bytes)")
                            
            except Exception as e:
                print(f"Error in Method 4: {e}")
            
            print(f"Total images extracted: {len(images)}")
            return images
            
        except Exception as e:
            print(f"Error extracting images: {e}")
            return []
    
    def save_image_from_drawing(self, image_obj, index):
        """Save image from drawing object"""
        try:
            # Get image data
            if hasattr(image_obj, '_data'):
                image_data = image_obj._data()
            elif hasattr(image_obj, 'ref'):
                # Try to get image from reference
                image_data = image_obj.ref
            else:
                print(f"Could not extract data from image {index}")
                return None
            
            # Determine file extension
            if isinstance(image_data, bytes):
                # Try to determine image type
                if image_data.startswith(b'\x89PNG'):
                    ext = 'png'
                elif image_data.startswith(b'\xff\xd8\xff'):
                    ext = 'jpg'
                elif image_data.startswith(b'GIF'):
                    ext = 'gif'
                else:
                    ext = 'png'  # Default
                
                filename = f"drawing_image_{index}.{ext}"
                filepath = os.path.join(self.images_dir, filename)
                
                with open(filepath, 'wb') as f:
                    f.write(image_data)
                
                return {
                    'filename': filename,
                    'path': filepath,
                    'source': f'drawing_{index}',
                    'size': len(image_data)
                }
            
        except Exception as e:
            print(f"Error saving drawing image {index}: {e}")
        
        return None
    
    def extract_image_from_relation(self, workbook, relation):
        """Extract image from workbook relation"""
        try:
            # Read the image data from the archive
            image_data = workbook._archive.read(relation.target)
            
            # Determine file extension from target path
            ext = relation.target.split('.')[-1].lower()
            if ext not in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
                ext = 'png'
            
            filename = f"relation_image_{relation.rId}.{ext}"
            filepath = os.path.join(self.images_dir, filename)
            
            with open(filepath, 'wb') as f:
                f.write(image_data)
            
            return {
                'filename': filename,
                'path': filepath,
                'source': relation.target,
                'size': len(image_data)
            }
            
        except Exception as e:
            print(f"Error extracting image from relation: {e}")
            return None
    
    def analyze_excel_structure(self):
        """Analyze Excel structure to understand layout"""
        try:
            # Load with pandas to understand data structure
            df = pd.read_excel(self.excel_path, header=None)
            print(f"Excel structure analysis:")
            print(f"Shape: {df.shape}")
            
            # Look for patterns that might indicate image locations
            image_cells = []
            for row in range(min(20, df.shape[0])):
                for col in range(min(20, df.shape[1])):
                    cell_value = df.iloc[row, col]
                    if pd.notna(cell_value):
                        cell_str = str(cell_value)
                        if any(keyword in cell_str.upper() for keyword in ['OGBE', 'OYEKU', 'IWORI', 'ODI']):
                            image_cells.append((row, col, cell_str))
            
            print(f"Found {len(image_cells)} cells with Odu names")
            for row, col, content in image_cells[:10]:
                print(f"  Cell ({row}, {col}): {content[:50]}...")
            
            return image_cells
            
        except Exception as e:
            print(f"Error analyzing Excel structure: {e}")
            return []
    
    def create_image_manifest(self, images, odu_cells):
        """Create manifest mapping images to Odu"""
        manifest = {
            "metadata": {
                "source": "256 Odu Excel Matrix Images",
                "extracted_at": datetime.now().isoformat(),
                "total_images": len(images),
                "total_odu_cells": len(odu_cells)
            },
            "images": images,
            "odu_cells": odu_cells,
            "mapping": []
        }
        
        # Try to create mappings between images and Odu
        if len(images) > 0 and len(odu_cells) > 0:
            print("Creating image-to-Odu mappings...")
            
            # Simple mapping: distribute images across Odu cells
            for i, (row, col, content) in enumerate(odu_cells):
                if i < len(images):
                    mapping = {
                        "odu_name": content.split('\n')[1] if '\n' in content else content,
                        "odu_cell": f"({row}, {col})",
                        "image": images[i],
                        "odu_number": i + 1
                    }
                    manifest["mapping"].append(mapping)
        
        # Save manifest
        manifest_path = os.path.join(self.output_dir, "excel_images_manifest.json")
        with open(manifest_path, 'w', encoding='utf-8') as f:
            json.dump(manifest, f, indent=2, ensure_ascii=False)
        
        print(f"Created image manifest: {manifest_path}")
        return manifest
    
    def update_odu_card_system(self, manifest):
        """Update the Odu card system to use extracted images"""
        if not manifest["images"]:
            print("No images found to integrate")
            return
        
        # Update the authentic Odu data to include image references
        try:
            with open("data/authentic_256_odu_excel.json", 'r', encoding='utf-8') as f:
                odu_data = json.load(f)
            
            # Update Odu entries with image paths
            for i, odu in enumerate(odu_data):
                if i < len(manifest["images"]):
                    image_info = manifest["images"][i]
                    odu["image_path"] = f"/static/odu_cards/images/{image_info['filename']}"
                    odu["has_authentic_image"] = True
                else:
                    odu["has_authentic_image"] = False
            
            # Save updated data
            with open("data/authentic_256_odu_excel.json", 'w', encoding='utf-8') as f:
                json.dump(odu_data, f, indent=2, ensure_ascii=False)
            
            # Update TypeScript data
            ts_content = f"""// Auto-generated 256 Odu data with extracted Excel images
// Generated: {datetime.now().isoformat()}

export const authentic256OduData = {json.dumps(odu_data, indent=2, ensure_ascii=False)};

export const extractedImages = {json.dumps(manifest["images"], indent=2, ensure_ascii=False)};

export function getOduByNumber(number: number) {{
  return authentic256OduData.find(odu => odu.number === number);
}}

export function getOduWithImage(number: number) {{
  const odu = getOduByNumber(number);
  return odu && odu.has_authentic_image ? odu : null;
}}
"""
            
            with open("server/authentic-odu-data.ts", 'w', encoding='utf-8') as f:
                f.write(ts_content)
            
            print(f"Updated {len(odu_data)} Odu entries with image references")
            
        except Exception as e:
            print(f"Error updating Odu card system: {e}")
    
    def run_complete_extraction(self):
        """Run complete image extraction process"""
        print("🖼️ Starting Excel Image Extraction...")
        
        # Analyze Excel structure
        odu_cells = self.analyze_excel_structure()
        
        # Extract images
        images = self.extract_images_from_excel()
        
        # Create manifest
        manifest = self.create_image_manifest(images, odu_cells)
        
        # Update Odu card system
        self.update_odu_card_system(manifest)
        
        print(f"""
✅ Excel Image Extraction Complete!

📊 Results:
- Images Extracted: {len(images)}
- Odu Cells Found: {len(odu_cells)}
- Mappings Created: {len(manifest.get('mapping', []))}

📁 Files:
- Images saved to: {self.images_dir}
- Manifest: {self.output_dir}/excel_images_manifest.json
- Updated Odu data with image references

🎯 Images ready for Odu card display!
""")
        
        return len(images) > 0

def main():
    extractor = ExcelImageExtractor()
    success = extractor.run_complete_extraction()
    
    if success:
        print("🎉 Image extraction successful!")
    else:
        print("⚠️ No images found in Excel file - may need manual extraction")

if __name__ == "__main__":
    main()